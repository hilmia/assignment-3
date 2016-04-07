import os
import wikipedia
from bs4 import BeautifulSoup 
import openpyxl

dictionary = dict()
final_list = {}
final_commits_list = {}
user_average = {}

wb = openpyxl.Workbook()
wb = openpyxl.load_workbook(filename ='a3WB.xlsx')

def find_average(country):
	sheets = wb.sheetnames
	sheet_avg = wb[sheets[0]]
	for i in range(2,117):
		row_num = i
		country_name_C = "A" + str(row_num)
		average_C = "D" + str(row_num)
		if country==sheet_avg[country_name_C].value:
			return float(sheet_avg[average_C].value)


def normalize_by_average(i):
	sheets = wb.sheetnames
	sheet = wb[sheets[2]]
	country_name = sheet["C" + str(i)].value
	average = dictionary.get(country_name)
	if(average == None):
		print country_name + " Not in Hash"
		average = find_average(country_name)
		dictionary[country_name] = average
	commits = sheet["D" + str(i)].value
	normalized_commits = float(commits/average)
	normalized_location = "E" + str(i)
	sheet[normalized_location].value = normalized_commits
	return normalized_commits

def normalize():
	sheets = wb.sheetnames
	sheet = wb[sheets[2]]
	current_project = "ActionBarSherlock"
	project_dictionary = dict()
	total_commits = 0
	users = 0
	for i in range(2,12613):
		print i
		country_name = sheet["C" + str(i)].value
		project_name = sheet["B" + str(i)].value
		if current_project == project_name:
			c = sheet["D"+str(i)].value
			total_commits = total_commits + c
			if(project_dictionary.get(country_name) == None):
				project_dictionary[country_name] = c
			else:
				project_dictionary[country_name] = project_dictionary[country_name] + c
			users = users + 1
		else:
			final_list[project_name] = project_dictionary
			current_project = project_name
			user_average[project_name] = float(total_commits)/users
			final_commits_list[project_name] = float(total_commits)
			users = 1
			total_commits = 0
			project_dictionary = dict()
			c = sheet["D"+str(i)].value
			total_commits = total_commits + c
			if(project_dictionary.get(country_name) == None):
				project_dictionary[country_name] = c
			else:
				project_dictionary[country_name] = project_dictionary[country_name] + c


def findMaxAndPrint():
	sheets = wb.sheetnames
	sheet = wb[sheets[3]]
	row = 2
	for keys in final_list:
		maxx = 0
		country = ""
		for t in final_list[keys]:
			if final_list[keys][t] > maxx:
				maxx = final_list[keys][t]
				country = t
		average = find_average(country)
		total_commits = final_commits_list[keys]
		normalized_percent = float(float(maxx)/float(total_commits)) * 100
		final_combination = float(normalized_percent)/float(average)
		average_user = user_average[keys]
		print keys
		print average_user
		#print final_combination
		#print average_user
		sheet["A"+str(row)] = final_combination
		sheet["B"+str(row)] = average_user
		row = row + 1







			

normalize()
findMaxAndPrint()
wb.save('temp.xlsx')







